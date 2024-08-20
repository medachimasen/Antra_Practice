import openpyxl

# create a new workbook
calc_book = openpyxl.Workbook()

# rename the default sheet
cursheet = calc_book.active
cursheet.title = "Calculator"

# put text labels in
cursheet.cell(1,1).value = "First number ==>"
cursheet.cell(2,1).value = "Second number ==>"

# add the numbers
cursheet.cell(1,2).value = 13
cursheet.cell(2,2).value = 8

# formula
cursheet.cell(3,2).value = "=B1+B2"

# save
calc_book.save(r"C:\\Users\Ben\Documents\Antra Homework\\calculator.xlsx")
calc_book.close()
